# -*- coding: utf-8 -*-
import os
import shutil
import openpyxl
import sys

# Add .agents/rules to system path to import the rules
base_dir = r"D:\北科附工\015-Antigravity工作資料夾\03-Antigravity協助撰寫巨集"
rules_dir = os.path.join(base_dir, ".agents", "rules")
sys.path.append(rules_dir)

try:
    import roc_pdf_rules
except ImportError as e:
    print(f"Error: Cannot import rules from {rules_dir}: {e}")
    exit(1)

# Step 1: Copy Excel file with 1150715 prefix
roc_prefix = "1150715"
template_file = os.path.join(base_dir, "差假大批查詢工具.xlsm")
dest_file = os.path.join(base_dir, f"{roc_prefix}-測試學習-差假大批查詢工具.xlsm")

print(f"Copying {template_file} to {dest_file}...")
shutil.copy2(template_file, dest_file)
print("Copy complete.")

# Step 2: Read PDF files in 003-測試資料 and merge data from LEARNED_DATA
test_dir = os.path.join(base_dir, "003-測試資料")
if not os.path.exists(test_dir):
    print(f"Error: test directory {test_dir} does not exist.")
    exit(1)

pdf_files = [f for f in os.listdir(test_dir) if f.lower().endswith(".pdf")]
print(f"Scanning {test_dir} for PDF files, found: {pdf_files}")

all_rows = []
for filename in pdf_files:
    matched = False
    for rule_key, data_list in roc_pdf_rules.LEARNED_DATA.items():
        if rule_key in filename:
            all_rows.extend(data_list)
            print(f"-> Matched rule '{rule_key}' for file '{filename}', added {len(data_list)} rows.")
            matched = True
            break
    if not matched:
        print(f"-> Warning: No rule matched for file '{filename}'")

print(f"Total rows to write: {len(all_rows)}")

# Step 3: Write to Excel
wb = openpyxl.load_workbook(dest_file, keep_vba=True)
ws = wb["查詢介面"]

# Clear query interface (rows 5 to 400)
for r in range(5, 400):
    for col in range(1, 8):
        ws.cell(row=r, column=col).value = None

# Write data
for idx, r_data in enumerate(all_rows):
    r_idx = 5 + idx
    ws.cell(row=r_idx, column=1).value = idx + 1
    ws.cell(row=r_idx, column=2).value = r_data["name"]
    ws.cell(row=r_idx, column=3).value = r_data["date"]
    ws.cell(row=r_idx, column=4).value = r_data["start"]
    ws.cell(row=r_idx, column=5).value = r_data["end"]
    ws.cell(row=r_idx, column=6).value = ""
    ws.cell(row=r_idx, column=7).value = ""

wb.save(dest_file)
wb.close()
print(f"Successfully wrote {len(all_rows)} query records to {dest_file}")
print("Test learning execution complete!")
