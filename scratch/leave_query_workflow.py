"""
#請假查詢 工作流程腳本
用法：python scratch/leave_query_workflow.py

功能：
1. 複製「差假大批查詢工具.xlsm」並加上今日民國日期前綴，但絕對不更動、不修改、不刪除「教職員名單」工作表中的資料！
2. 使用 winsdk Windows Media OCR 讀取「002-待作業檔案」中所有 PDF（支援掃描型/圖片型 PDF）
3. 解析姓名、日期、時間並填入複製後的 xlsm 查詢介面
"""
import sys, io, os, re, shutil, subprocess, asyncio
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from datetime import datetime, date

# 嘗試載入 winsdk 模組
try:
    from winsdk.windows.storage import StorageFile
    from winsdk.windows.graphics.imaging import BitmapDecoder
    from winsdk.windows.media.ocr import OcrEngine
    from winsdk.windows.globalization import Language
except ImportError:
    print("❌ 缺少 winsdk 模組，請先執行 pip install winsdk 安裝。")
    sys.exit(1)

import pypdfium2 as pdfium

# ── 路徑設定 ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    # 若為打包後的 exe 執行環境
    PROJECT_DIR = os.path.dirname(sys.executable)
else:
    # 原始開發環境
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    PROJECT_DIR = os.path.dirname(SCRIPT_DIR)

TEMPLATE     = os.path.join(PROJECT_DIR, "差假大批查詢工具.xlsm")
PDF_INBOX    = os.path.join(PROJECT_DIR, "002-待作業檔案")

ROC_OFFSET   = 1911

# ── 載入教職員名單 ─────────────────────────────────────────
def load_staff_names() -> list[str]:
    import openpyxl
    names = []
    try:
        wb = openpyxl.load_workbook(TEMPLATE, read_only=True)
        if "教職員名單" in wb.sheetnames:
            ws = wb["教職員名單"]
            for r in range(1, 1000):
                val = ws.cell(r, 1).value
                if val:
                    val_str = str(val).strip()
                    if val_str and val_str != "姓名":
                        names.append(val_str)
    except Exception as e:
        print("⚠️ 無法從範本載入教職員名單：", e)
    return names

# ── 民國日期工具 ──────────────────────────────────────────
def today_roc_str() -> str:
    today = date.today()
    roc_year = today.year - ROC_OFFSET
    return f"{roc_year:03d}{today.month:02d}{today.day:02d}"

def resolve_roc_year(school_year_str: str, month: int) -> int:
    """根據台灣學年度與月份推算民國年 (8月至翌年7月為同一學年度)"""
    try:
        school_year = int(school_year_str)
        return school_year if month >= 8 else school_year + 1
    except ValueError:
        return date.today().year - ROC_OFFSET

# ── 本地 Windows OCR 核心邏輯 ────────────────────────────────
async def ocr_page_async(png_path: str) -> str:
    abs_path = os.path.abspath(png_path)
    file = await StorageFile.get_file_from_path_async(abs_path)
    stream = await file.open_read_async()
    decoder = await BitmapDecoder.create_async(stream)
    software_bitmap = await decoder.get_software_bitmap_async()
    
    lang = Language("zh-Hant-TW")
    if OcrEngine.is_language_supported(lang):
        engine = OcrEngine.try_create_from_language(lang)
    else:
        engine = OcrEngine.try_create_from_user_profile_languages()
        
    result = await engine.recognize_async(software_bitmap)
    return result.text

async def ocr_pdf_file_async(pdf_path: str) -> str:
    doc = pdfium.PdfDocument(pdf_path)
    full_text = []
    print(f"  📖 正在對 {os.path.basename(pdf_path)} 進行本機 OCR 辨識 (頁數: {len(doc)})...")
    
    for idx, page in enumerate(doc):
        bitmap = page.render(scale=3) # 解析度 3 倍提升 OCR 準確度
        pil_img = bitmap.to_pil()
        temp_png = f"temp_ocr_page_{idx}.png"
        pil_img.save(temp_png)
        
        try:
            text = await ocr_page_async(temp_png)
            full_text.append(text)
        finally:
            if os.path.exists(temp_png):
                os.remove(temp_png)
                
    return "\n".join(full_text)

# ── 空間與文字標準化 ───────────────────────────────────────
def normalize_text(text: str) -> str:
    """消除中文與數字間的多餘空格，保留分行資訊方便逐行解析"""
    lines = []
    for line in text.splitlines():
        line_clean = re.sub(r'[ \t]+', '', line)
        if line_clean:
            lines.append(line_clean)
    return "\n".join(lines)

# ── 解析提取邏輯 ─────────────────────────────────────────
def parse_extracted_data(text: str, filename: str, staff_names: list[str]) -> list[dict]:
    rows = []
    normalized = normalize_text(text)
    
    # 1. 識別學年度 (用於推算沒有寫年份的月日，如 3/12)
    school_year = "114" # 預設 fallback
    m_sy = re.search(r'(\d{3})學年度', normalized)
    if m_sy:
        school_year = m_sy.group(1)
        
    # 2. 識別姓名 (優先以檔名匹配，次以教職員名單匹配)
    name = None
    for staff in staff_names:
        if staff in filename:
            name = staff
            break
            
    # 3. 根據文件類型執行特定解析
    # 類型 A: 職場巡輔交通費申請表 (格式多為 3/12、4/2、5/21)
    if "職場巡輔" in normalized or "巡輔" in normalized or "交通費" in normalized:
        if not name:
            for staff in staff_names:
                if staff in normalized:
                    name = staff
                    break
        
        if not name:
            print(f"  ⚠️  無法在檔名或內文中匹配到教職員名單中的姓名，跳過此檔案")
            return []
            
        print(f"  👤 姓名：{name} (職場巡輔申請表)")
        
        seen_dates = set()
        date_patterns = [
            r'(\d{1,2})/(\d{1,2})',
            r'(\d{1,2})月(\d{1,2})日'
        ]
        
        for pat in date_patterns:
            for m in re.finditer(pat, normalized):
                try:
                    mo = int(m.group(1))
                    d = int(m.group(2))
                    if mo < 1 or mo > 12 or d < 1 or d > 31:
                        continue
                    
                    y = resolve_roc_year(school_year, mo)
                    roc_date = f"{y:03d}{mo:02d}{d:02d}"
                    
                    if roc_date not in seen_dates:
                        seen_dates.add(roc_date)
                        rows.append({"name": name, "date_roc": roc_date, "start": "0800", "end": "1200"})
                        rows.append({"name": name, "date_roc": roc_date, "start": "1300", "end": "1700"})
                except Exception:
                    pass
                    
        rows = sorted(rows, key=lambda x: (x["date_roc"], x["start"]))
        print(f"  📅 職場巡輔：提取出 {len(seen_dates)} 個日期，共 {len(rows)} 筆查詢時段")
        
    # 類型 B: 印領清冊/鐘點費明細 (逐列或整頁比對日期與教師)
    else:
        print("  ℹ️  使用清冊比對策略...")
        # 3.1 找出整個 PDF 中出現過哪些教師的姓名
        teachers_in_doc = []
        for staff in staff_names:
            if staff in normalized:
                teachers_in_doc.append(staff)
        
        # 3.2 如果在整個文件中只匹配到唯一一位教師，則將文件中所有找到的日期都歸屬於他
        if len(teachers_in_doc) == 1:
            name = teachers_in_doc[0]
            m_dates = re.finditer(r'(\d{2,3})年(\d{1,2})月(\d{1,2})日', normalized)
            seen_dates = set()
            for m in m_dates:
                try:
                    y = int(m.group(1))
                    mo = int(m.group(2))
                    d = int(m.group(3))
                    roc_date = f"{y:03d}{mo:02d}{d:02d}"
                    if roc_date not in seen_dates:
                        seen_dates.add(roc_date)
                        rows.append({"name": name, "date_roc": roc_date, "start": "0800", "end": "1500"})
                        print(f"  👤 匹配教師：{name} (唯一教師) | 📅 授課日期：{roc_date} (預設 0800 ~ 1500)")
                except Exception:
                    pass
        else:
            # 如果有多個教師或沒有，則使用嚴格的「同一行」匹配日期與姓名
            lines = normalized.splitlines()
            for line in lines:
                m_date = re.search(r'(\d{2,3})年(\d{1,2})月(\d{1,2})日', line)
                if m_date:
                    try:
                        y = int(m_date.group(1))
                        mo = int(m_date.group(2))
                        d = int(m_date.group(3))
                        roc_date = f"{y:03d}{mo:02d}{d:02d}"
                        
                        matched_teacher = None
                        for staff in staff_names:
                            if staff in line:
                                matched_teacher = staff
                                break
                                
                        if matched_teacher:
                            rows.append({"name": matched_teacher, "date_roc": roc_date, "start": "0800", "end": "1500"})
                            print(f"  👤 匹配教師：{matched_teacher} | 📅 授課日期：{roc_date} (預設 0800 ~ 1500)")
                    except Exception:
                        pass
                        
        if not rows:
            print(f"  ⚠️  無法在清冊中匹配到任何日期與教師資料列")
            
    return rows

# ── Excel 寫入 ────────────────────────────────────────────
def write_to_excel(xlsm_path: str, rows: list[dict]):
    import win32com.client
    
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    
    try:
        print(f"\n✍️ 正在將資料寫入 Excel...")
        wb = excel.Workbooks.Open(xlsm_path)
        ws_query = wb.Sheets("查詢介面")
        
        # 1. 清除「查詢介面」舊有查詢資料 (B5 以後)，但絕對不碰其他工作表！
        last_used = ws_query.Cells(ws_query.Rows.Count, 2).End(-4162).Row # xlUp = -4162
        if last_used >= 5:
            ws_query.Range(f"A5:G{last_used}").ClearContents()
            ws_query.Range(f"A5:G{last_used}").Interior.ColorIndex = -4142 # xlNone
            ws_query.Range(f"A5:G{last_used}").Font.ColorIndex = -4105      # xlAutomatic
            
        # 2. 寫入新提取的查詢資料
        for i, row in enumerate(rows):
            r = 5 + i
            ws_query.Cells(r, 1).Value = i + 1              # 序號
            ws_query.Cells(r, 2).Value = row["name"]         # 姓名
            ws_query.Cells(r, 3).NumberFormat = "@"
            ws_query.Cells(r, 3).Value = row["date_roc"]     # 日期 (字串)
            ws_query.Cells(r, 4).NumberFormat = "@"
            ws_query.Cells(r, 4).Value = row["start"]        # 開始時間
            ws_query.Cells(r, 5).NumberFormat = "@"
            ws_query.Cells(r, 5).Value = row["end"]          # 結束時間
            
        wb.Save()
        print(f"🎉 成功寫入 {len(rows)} 筆資料到：{os.path.basename(xlsm_path)}")
        
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass
        # 強制關閉 Excel 背景進程釋放檔案鎖
        subprocess.run(
            ["powershell", "-Command",
             "Get-Process EXCEL -ErrorAction SilentlyContinue | Stop-Process -Force"],
            capture_output=True
        )

# ── 主流程 ────────────────────────────────────────────────
async def main_async():
    print("=" * 60)
    print("  #請假查詢 工作流程啟動 (winsdk OCR 離線版)")
    print("=" * 60)
    
    # 0. 載入教職員名單
    staff_names = load_staff_names()
    print(f"👥 已載入 {len(staff_names)} 位教職員名單作比對基礎")
    if not staff_names:
        print("❌ 教職員名單為空，無法進行比對識別，請先在「差假大批查詢工具.xlsm」的教職員名單分頁中建置名單。")
        return
        
    # 1. 取得今日民國日期
    roc_today = today_roc_str()
    print(f"📅 今日民國日期：{roc_today}")
    
    # 2. 複製活頁簿 (原地保留所有工作表與名單)
    if not os.path.exists(TEMPLATE):
        print(f"❌ 找不到範本檔案：{TEMPLATE}")
        return
        
    dest_name = f"{roc_today}-差假大批查詢工具.xlsm"
    dest_path = os.path.join(PROJECT_DIR, dest_name)
    
    # 如果已存在，直接複製覆蓋，這會帶有原範本所有的工作表格式
    shutil.copy2(TEMPLATE, dest_path)
    print(f"✅ 成功建立今日查詢工作簿：{dest_name}")
    
    # 3. 掃描待作業 PDF
    if not os.path.exists(PDF_INBOX):
        print(f"❌ 找不到待作業資料夾：{PDF_INBOX}")
        return
        
    pdf_files = [
        os.path.join(PDF_INBOX, f)
        for f in sorted(os.listdir(PDF_INBOX))
        if f.lower().endswith(".pdf")
    ]
    
    if not pdf_files:
        print(f"\n📂 「002-待作業檔案」資料夾目前沒有 PDF 檔案。")
        print("   請放檔案後再執行 #請假查詢。")
        return
        
    print(f"\n📂 找到 {len(pdf_files)} 個待處理 PDF 檔案：")
    for p in pdf_files:
        print(f"  - {os.path.basename(p)}")
        
    # 4. 對每個 PDF 進行 OCR 辨識與解析
    all_rows = []
    for pdf_path in pdf_files:
        try:
            text = await ocr_pdf_file_async(pdf_path)
            rows = parse_extracted_data(text, os.path.basename(pdf_path), staff_names)
            all_rows.extend(rows)
        except Exception as e:
            print(f"  ❌ {os.path.basename(pdf_path)} 辨識或解析失敗：{e}")
            import traceback
            traceback.print_exc()
            
    if not all_rows:
        print("\n⚠️ 檔案解析後無有效時段資料，請確認 PDF 內容。")
        return
        
    print(f"\n📊 統計：共從 {len(pdf_files)} 個 PDF 中提取出 {len(all_rows)} 筆查詢時段")
    
    # 5. 寫入複製的 Excel 檔案中
    write_to_excel(dest_path, all_rows)
    
    print("\n" + "=" * 60)
    print("  🎉 #請假查詢 任務已完成！")
    print("=" * 60)
    print(f"\n  📁 今日工作檔案：{dest_name}")
    print(f"  📄 已辨識 PDF 數：{len(pdf_files)} 個")
    print(f"  📝 填入查詢筆數：{len(all_rows)} 筆")
    print("\n  💡 教職員名單保護：")
    print("     您的「教職員名單」工作表完全完好如初，絕無任何刪除或修改。")
    print("\n  ➡️  後續操作：")
    print("     1. 請開啟檔案")
    print("     2. 點選「1. 匯入請假明細」匯入差勤 XLS 檔")
    print("     3. 點選「2. 執行大批查詢」進行比對與著色")

if __name__ == "__main__":
    try:
        asyncio.run(main_async())
    except Exception as e:
        print(f"\n❌ 發生未預期錯誤：{e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\n請按 Enter 鍵結束並關閉視窗...")
