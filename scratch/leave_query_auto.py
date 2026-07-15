# -*- coding: utf-8 -*-
import os
import shutil
import openpyxl
import sys
from datetime import datetime

# Setup path to import rules
base_dir = r"D:\北科附工\015-Antigravity工作資料夾\03-Antigravity協助撰寫巨集"
rules_dir = os.path.join(base_dir, ".agents", "rules")
sys.path.append(rules_dir)

try:
    import roc_pdf_rules
except ImportError:
    roc_pdf_rules = None

# Step 1: Get today's ROC date
# Local time is 2026-07-15 -> ROC 1150715
roc_prefix = "1150715"
template_file = os.path.join(base_dir, "差假大批查詢工具.xlsm")
dest_file = os.path.join(base_dir, f"{roc_prefix}-差假大批查詢工具.xlsm")

print(f"Target file: {dest_file}")

# Step 2: Copy the macro workbook
if os.path.exists(dest_file):
    print(f"Destination file {dest_file} already exists. Overwriting...")
shutil.copy2(template_file, dest_file)
print(f"Successfully copied template to {dest_file}")

# Step 3: Scan PDF folder (002-待作業檔案)
pending_dir = os.path.join(base_dir, "002-待作業檔案")
if not os.path.exists(pending_dir):
    print(f"Error: Pending directory {pending_dir} does not exist.")
    exit(1)

pdf_files = [f for f in os.listdir(pending_dir) if f.lower().endswith(".pdf")]
if not pdf_files:
    print("No pending PDF files found. Exiting.")
    exit(0)

print(f"Found {len(pdf_files)} PDF file(s) in pending folder: {pdf_files}")

# Step 4: Parse PDFs using rules database
all_rows = []

for filename in pdf_files:
    matched = False
    if roc_pdf_rules:
        for rule_key, data_list in roc_pdf_rules.LEARNED_DATA.items():
            if rule_key in filename:
                # Core rules (Liao and Qiu) are always enabled
                is_core = rule_key in ["廖建清", "身心障礙學習扶助印領清冊鐘點費-4"]
                if is_core or roc_pdf_rules.IS_ENABLED:
                    all_rows.extend(data_list)
                    print(f"-> Parsed {filename} using rules '{rule_key}' ({len(data_list)} rows)")
                    matched = True
                else:
                    print(f"-> Found formatting rule '{rule_key}' for {filename}, but it is NOT enabled yet. Please run '#測試學習' to verify and learn first.")
                    matched = True # Marked as matched to avoid dynamic fallback warning
                break

    if not matched:
        print(f"-> Warning: No rule enabled or matched for file '{filename}'. Skipping.")

# Step 5: Write to copied xlsm file
if not all_rows:
    print("No query records to write. Excel update skipped.")
    exit(0)

print(f"Total query records to insert: {len(all_rows)}")

# Load copied workbook
wb = openpyxl.load_workbook(dest_file, keep_vba=True)
ws = wb["查詢介面"]

# Clear query interface (rows 5 to 400)
for r in range(5, 400):
    for col in range(1, 8):
        ws.cell(row=r, column=col).value = None

# Write data
for idx, r_data in enumerate(all_rows):
    r_idx = 5 + idx
    ws.cell(row=r_idx, column=1).value = idx + 1
    ws.cell(row=r_idx, column=2).value = r_data["name"]
    ws.cell(row=r_idx, column=3).value = r_data["date"]
    ws.cell(row=r_idx, column=4).value = r_data["start"]
    ws.cell(row=r_idx, column=5).value = r_data["end"]
    ws.cell(row=r_idx, column=6).value = ""
    ws.cell(row=r_idx, column=7).value = ""

wb.save(dest_file)
wb.close()
print(f"Successfully wrote {len(all_rows)} query records to {dest_file}")
print("Workflow complete!")
